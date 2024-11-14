import openpyxl
from openpyxl.styles import Font
import os
import telebot
import sqlite3
import re
import csv 
from datetime import datetime
from telebot.types import ForceReply, InlineKeyboardMarkup, InlineKeyboardButton
#from config import * #importar config
from pacientesBD import * #Importamos pacientes, donde se crea la base de datos

#Instanciar el bot
bot = telebot.TeleBot("7734067796:AAF_2wvdRTR9KQi6jYPOcf-PKyYfSx6eyGk")

ALLOWED_USERS = [1130744210, 1163297608]
seguimiento_data = {}
temp_data = {}
global usuario_id

################## COMANDOS ##################
@bot.message_handler(commands=["id"])
def cmd_id(message):
    print(message.from_user.id)

#Responde a comando /start
@bot.message_handler(commands= ["iniciar"])
def cmd_start(message):
    #Dar bienvenida al usuario
    bot.reply_to(message, "Hola! 1.\n Usa el comando /seguimiento para comenzar el seguimiento de tu paciente \n 2. Usa el comando /buscar_seguimiento para ver tus seguimientos")

#Registrar Pacientes desde Admin, comando /registrar
@bot.message_handler(commands=["registrar"])
def cmd_registrar(message):

    if message.from_user.id not in ALLOWED_USERS:
     bot.reply_to(message, "No tienes permiso para usar este comando.")
     return

    global usuario_id
    usuario_id = message.from_user.id
    temp_data[usuario_id] = {}

    bot.reply_to(message, "Ingresa el folio del paciente: ")
    bot.register_next_step_handler(message, obtener_folio)

#Seguimiento de pacientes, comando /seguimiento
@bot.message_handler(commands=["seguimiento"])
def cmd_seguimiento(message):
    bot.send_message(message.chat.id, "Ingrese folio de su paciente")
    bot.register_next_step_handler(message, validar_folio)

#Buscar paciente, comando /buscar_paciente
@bot.message_handler(commands=['buscar_paciente'])
def cmd_buscar_paciente(message):
    if message.from_user.id not in ALLOWED_USERS:
     bot.reply_to(message, "No tienes permiso para usar este comando.")
     return
    
    bot.reply_to(message, "Ingresa el folio del paciente para buscarlo.")
    bot.register_next_step_handler(message, buscar_paciente)

#Buscar paciente, comando /buscar_seguimiento
@bot.message_handler(commands=['buscar_seguimiento'])
def cmd_buscar_seguimiento(message):
    buscar_seguimiento(message)

#Exportar Seguimientos en excel
@bot.message_handler(commands=['exportar'])
def exportar_excel(message):
    conn = sqlite3.connect('pacientes.db')
    cursor = conn.cursor()

    # Obtener todos los datos de la tabla
    cursor.execute('SELECT * FROM seguimientos')
    resultados = cursor.fetchall()
    conn.close()

    # Crear un nuevo archivo Excel
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Seguimientos"

    # Escribir encabezados
    encabezados = ['Indice','Folio', 'Fecha', 'Hora', 'Temperatura', 'Vomitos', 'frecuencia_vomitos', 'problemas_respiracion', 'dolor_corporal', 'zona_dolor', 'intensidad_dolor']  # Cabeceras
    for col, encabezado in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=col).value = encabezado
        hoja.cell(row=1, column=col).font = Font(bold=True)  # Encabezados en negrita

    # Escribir los datos en el archivo
    for fila, datos in enumerate(resultados, start=2):
        for col, valor in enumerate(datos, start=1):
            hoja.cell(row=fila, column=col).value = valor

    # Guardar el archivo temporalmente
    archivo_excel = "reporte_pacientes.xlsx"
    workbook.save(archivo_excel)

    # Enviar el archivo al usuario
    with open(archivo_excel, 'rb') as file:
        bot.send_document(message.chat.id, file)

    # Eliminar el archivo después de enviarlo
    os.remove(archivo_excel)

################## REGISTRO DE PACIENTES ##################

#Funcion para obtener el folio
def obtener_folio(message):
    usuario_id = message.from_user.id
    temp_data[usuario_id]['folio'] = message.text

    conn = sqlite3.connect('pacientes.db')
    cursor = conn.cursor()
    cursor.execute('SELECT folio FROM pacientes WHERE folio = ?', (temp_data[usuario_id]['folio'],))

    if cursor.fetchone():
        bot.reply_to(message, "El paciente ya está registrado.")
        conn.close()
        bot.reply_to(message, "Ingresa el folio del paciente")
        bot.register_next_step_handler(message, obtener_folio)
        return
    else:
      bot.reply_to(message, "Ingresa el nombre del paciente")
      bot.register_next_step_handler(message, obtener_nombre)

#Funcion para obtener el nombre
def obtener_nombre(message):
    usuario_id = message.from_user.id
    nombre = message.text.strip()
    
    # Validar formato del nombre
    if not re.match(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ ]+$", nombre):
        bot.reply_to(message, "El nombre solo puede contener letras y espacios. Intenta nuevamente.")
        bot.reply_to(message, "Ingresa el nombre del paciente")
        bot.register_next_step_handler(message, obtener_nombre)
    elif True:
        temp_data[usuario_id]['nombre'] = message.text

        bot.reply_to(message, "Ingresa el apellido paterno")
        bot.register_next_step_handler(message, obtener_apellido_paterno)

#Funcion para obtener el apellido paterno
def obtener_apellido_paterno(message):
    usuario_id = message.from_user.id
    apellido = message.text.strip()

    # Validar formato del nombre
    if not re.match(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ ]+$", apellido):
        bot.reply_to(message, "El apellido solo puede contener letras y espacios. Intenta nuevamente.")
        bot.reply_to(message, "Ingresa el apellido paterno del paciente")
        bot.register_next_step_handler(message, obtener_apellido_paterno)
    elif True:
        temp_data[usuario_id]['apellido_paterno'] = message.text
        bot.reply_to(message, "Ingresa apellido materno")
        bot.register_next_step_handler(message, obtener_apellido_materno)

#Funcion para obtener el apellido materno
def obtener_apellido_materno(message):
    usuario_id = message.from_user.id
    apellido = message.text.strip()

    # Validar formato del nombre
    if not re.match(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ ]+$", apellido):
        bot.reply_to(message, "El apellido solo puede contener letras y espacios. Intenta nuevamente.")
        bot.reply_to(message, "Ingresa el apellido paterno del paciente")
        bot.register_next_step_handler(message, obtener_apellido_paterno)
    elif True:
        temp_data[usuario_id]['apellido_materno'] = message.text
        bot.reply_to(message, "Ingresa la edad del paciente")
        bot.register_next_step_handler(message, obtener_edad)

#Funcion para obtener la edad
def obtener_edad(message):
    usuario_id = message.from_user.id
    edad = int(message.text)

    if edad >= 0 and edad <= 18:
        temp_data[usuario_id]['edad'] = int(message.text)
        bot.reply_to(message, "Ingresa el lugar de procedencia del paciente:")
        bot.register_next_step_handler(message, obtener_procedencia)
    else:
        bot.reply_to(message, "Por favor, ingresa un número válido para la edad(0-18).")
        bot.register_next_step_handler(message, obtener_edad)

##Funcion para obtener la procedencia del paciente
def obtener_procedencia(message):
    usuario_id = message.from_user.id
    temp_data[usuario_id]['lugar_procedencia'] = message.text

    bot.reply_to(message, "Por favor, ingresa un número de teléfono válido de 10 dígitos.")
    bot.register_next_step_handler(message, obtener_numero)

##Funcion para guardar el numero y llamar la funcion que guarda en la Base de datos sqlite
def obtener_numero(message):
    usuario_id = message.from_user.id

    if not re.fullmatch(r"^\d{10}$", message.text):
      bot.reply_to(message, "Por favor, ingresa un número de teléfono válido de 10 dígitos.")
      bot.register_next_step_handler(message, obtener_numero)
      return
    
    elif True:
        temp_data[usuario_id]['numero'] = int(message.text)

        guardar_en_db(usuario_id)

        bot.reply_to(message, "El registro ha sido completado exitosamente.")
        del temp_data[usuario_id]  # Eliminar los datos temporales

def guardar_en_db(usuario_id):
    paciente = temp_data[usuario_id]
    conn = sqlite3.connect('pacientes.db')
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO pacientes (folio, nombre, apellido_paterno, apellido_materno, edad, lugar_procedencia, numero)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (paciente['folio'], paciente['nombre'], paciente['apellido_paterno'], paciente['apellido_materno'],
        paciente['edad'], paciente['lugar_procedencia'], paciente['numero']))

    conn.commit()
    conn.close() # Cerramos la conexión

################## BUSCAR PACIENTES ##################

def buscar_paciente(message):
    print("entre")
    criterio = message.text.strip()

    if not criterio:
      bot.reply_to(message, "El criterio de búsqueda no puede estar vacío.")
      bot.reply_to(message, "Ingresa el folio del paciente para buscarlo.")
      bot.register_next_step_handler(message, buscar_paciente)

    conn = sqlite3.connect('pacientes.db')
    cursor = conn.cursor()

    # Buscar por folio o nombre
    cursor.execute('SELECT * FROM pacientes WHERE folio = ? OR nombre LIKE ?', (criterio, f"%{criterio}%"))
    resultados = cursor.fetchall()
    conn.close()

    if resultados:
        # Mostrar los resultados
        respuesta = "Resultados encontrados:\n\n"
        for paciente in resultados:
            respuesta += (f"Folio: {paciente[1]}\n"
                          f"Nombre: {paciente[2]}\n"
                          f"Apellido Paterno: {paciente[3]}\n"
                          f"Apellido Paterno: {paciente[4]}\n"
                          f"Edad: {paciente[5]}\n"
                          f"Procedencia: {paciente[6]}\n"
                          f"Número: {paciente[7]}\n\n")
        bot.reply_to(message, respuesta)
    else:
        bot.reply_to(message, "No se encontraron pacientes con ese criterio.")

################## SEGUIMIENTO ##################

#Obtener folio de paciente y avisar si se encuentra registrado
def validar_folio(message):
    folio = message.text
    global usuario_id
    #
    usuario_id = message.from_user.id
    paciente = verificar_paciente(folio)

    if paciente:
        # Inicializar temp_data para este usuario
        temp_data[usuario_id] = {'folio': folio}
        bot.send_message(message.chat.id, "Paciente encontrado. Iniciando seguimiento.")
        preguntar_temperatura(message)
    else:
        bot.send_message(message.chat.id, "Paciente no encontrado. Por favor verifique el folio.")

################## INICIA SEGUIMIENTO ##################

# Pregunta sobre la temperatura del paciente
def preguntar_temperatura(message):
    markup = InlineKeyboardMarkup()
    buttons = [
        InlineKeyboardButton("Menor a 36", callback_data="temp_menor_36"),
        InlineKeyboardButton("36 a 37", callback_data="temp_36_37"),
        InlineKeyboardButton("37 a 38", callback_data="temp_37_38"),
        InlineKeyboardButton("38 a 39", callback_data="temp_38_39"),
        InlineKeyboardButton("39 a 40", callback_data="temp_39_40"),
        InlineKeyboardButton("Mayor a 40", callback_data="temp_mayor_40")
    ]
    markup.add(*buttons)
    bot.send_message(message.chat.id, "Ingrese la temperatura de su paciente:", reply_markup=markup)

# Manejar respuestas de temperatura
@bot.callback_query_handler(func=lambda call: call.data.startswith("temp_"))
def respuesta_temperatura(call):
    usuario_id = call.from_user.id
    if call.data == "temp_menor_36":
            #temp_data[usuario_id]['temperatura'] = call.data.replace("temp_", "").replace("_", " ")
            temp_data[usuario_id]['temperatura'] = 36
            preguntar_vomitos(call.message)

    if call.data == "temp_mayor_40":
            #temp_data[usuario_id]['temperatura'] = call.data.replace("temp_", "").replace("_", " ")
            temp_data[usuario_id]['temperatura'] = 40
            preguntar_vomitos(call.message)

    if call.data == "temp_36_37":
        # Mostrar opciones específicas entre 36.1 y 36.9
        markup = InlineKeyboardMarkup()
        buttons = [InlineKeyboardButton(f"36.{i}", callback_data=f"temp36.{i}") for i in range(1, 10)]
        markup.add(*buttons)
        bot.edit_message_text("Elija una temperatura más específica:", call.message.chat.id, call.message.message_id, reply_markup=markup)
        print(call.data)

    if call.data == "temp_37_38":
        # Mostrar opciones específicas entre 36.1 y 36.9
        markup = InlineKeyboardMarkup()
        buttons = [InlineKeyboardButton(f"37.{i}", callback_data=f"temp37.{i}") for i in range(1, 10)]
        markup.add(*buttons)
        bot.edit_message_text("Elija una temperatura más específica:", call.message.chat.id, call.message.message_id, reply_markup=markup)
        print(call.data)

    if call.data == "temp_38_39":
        # Mostrar opciones específicas entre 36.1 y 36.9
        markup = InlineKeyboardMarkup()
        buttons = [InlineKeyboardButton(f"38.{i}", callback_data=f"temp38.{i}") for i in range(1, 10)]
        markup.add(*buttons)
        bot.edit_message_text("Elija una temperatura más específica:", call.message.chat.id, call.message.message_id, reply_markup=markup)
        print(call.data)

    if call.data == "temp_39_40":
        # Mostrar opciones específicas entre 36.1 y 36.9
        markup = InlineKeyboardMarkup()
        buttons = [InlineKeyboardButton(f"39.{i}", callback_data=f"temp39.{i}") for i in range(1, 10)]
        markup.add(*buttons)
        bot.edit_message_text("Elija una temperatura más específica:", call.message.chat.id, call.message.message_id, reply_markup=markup)
        print(call.data)


#Guardar Temperatura
@bot.callback_query_handler(func = lambda call: re.match(r'^temp(3[6-9]\.)', call.data) is not None)
def respuesta_temperatura_especifica(call):
    usuario_id = call.from_user.id
    temp_data[usuario_id]['temperatura'] = call.data.replace("temp", "")
    preguntar_vomitos(call.message)

#Preguntar Vomitos
def preguntar_vomitos(message):
    markup = InlineKeyboardMarkup()
    buttons = [InlineKeyboardButton("Si", callback_data="vomitos_si"), InlineKeyboardButton("No", callback_data="vomitos_no")]
    markup.add(*buttons)
    bot.send_message(message.chat.id, "¿Su paciente ha presentado vómitos?", reply_markup=markup)

#Respuestas sobre vómitos
@bot.callback_query_handler(func=lambda call: call.data.startswith("vomitos_"))
def respuesta_vomitos(call):
    usuario_id = call.from_user.id

    print(f"Datos de temp_data en respuesta_vomito: {temp_data}")

    if call.data == "vomitos_si":
        temp_data[usuario_id]['vomitos'] = "Si"
        markup = InlineKeyboardMarkup()
        buttons = [
        InlineKeyboardButton("1 vez por semana", callback_data="frec_1"),
        InlineKeyboardButton("2 veces por semana", callback_data="frec_2"),
        InlineKeyboardButton("3 o más veces por semana", callback_data="frec_3")
        ]
        markup.add(*buttons)
        bot.edit_message_text("¿Con qué frecuencia?", call.message.chat.id, call.message.message_id, reply_markup=markup)
    else:
        temp_data[usuario_id]['vomitos'] = "No"
        temp_data[usuario_id]['frecuencia_vomitos']  = '0'
        preguntar_respiracion(call.message)

#Frecuencia Vomitos
@bot.callback_query_handler(func=lambda call: call.data.startswith("frec_"))
def respuesta_frecuencia_vomitos(call):
    usuario_id = call.from_user.id

    print(f"Datos de temp_data en respuesta_temperatura: {temp_data}")

    temp_data[usuario_id]['frecuencia_vomitos'] = call.data.replace("frec_", "")
    preguntar_respiracion(call.message)

#Preguntar Respiracion
def preguntar_respiracion(message):
    markup = InlineKeyboardMarkup()
    buttons = [InlineKeyboardButton("Si", callback_data="resp_si"), InlineKeyboardButton("No", callback_data="resp_no")]
    markup.add(*buttons)
    bot.send_message(message.chat.id, "¿Su paciente presenta problemas para respirar?", reply_markup=markup)

#Respuestas problemas respiracion
@bot.callback_query_handler(func=lambda call: call.data.startswith("resp_"))
def respuesta_respiracion(call):
    usuario_id = call.from_user.id
    print(f"Datos de temp_data en respuesta_respiracion: {temp_data}")
    temp_data[usuario_id]['problemas_respiracion'] = call.data.replace("resp_", "")
    preguntar_dolor_corporal(call.message)

# Pregunta sobre dolor corporal
def preguntar_dolor_corporal(message):
    markup = InlineKeyboardMarkup()
    buttons = [InlineKeyboardButton("Si", callback_data="dolor_si"), InlineKeyboardButton("No", callback_data="dolor_no")]
    markup.add(*buttons)
    bot.send_message(message.chat.id, "¿Su paciente presenta dolor corporal?", reply_markup=markup)

#Respuestas dolor corporal
@bot.callback_query_handler(func=lambda call: call.data.startswith('dolor_'))
def respuesta_dolor_corporal(call):
    usuario_id = call.from_user.id
    print(f"Datos de temp_data en respuesta_dolor corporal: {temp_data}")
    if call.data == "dolor_si":
        temp_data[usuario_id]['dolor_corporal'] = "Si"
        preguntar_zona_dolor(call.message)
    else:
        temp_data[usuario_id]['dolor_corporal'] = "No"
        temp_data[usuario_id]['zona_dolor'] = "0"
        temp_data[usuario_id]['intensidad_dolor'] = "0"
        guardar_seguimiento(call.message)

#Preguntar Zona Dolor
def preguntar_zona_dolor(message):
    markup = InlineKeyboardMarkup()
    zonas = ["Cabeza", "Hombros", "Brazos", "Manos", "Pecho", "Estómago", "Piernas", "Pies", "Espalda"]
    buttons =[InlineKeyboardButton(zona, callback_data=f"zona_{zona}") for zona in zonas]
    markup.add(*buttons)
    bot.send_message(message.chat.id, "¿En qué zona del cuerpo se presenta el dolor?", reply_markup=markup)

# Respuestas sobre la zona del dolor
@bot.callback_query_handler(func=lambda call: call.data.startswith("zona_"))
def respuesta_zona_dolor(call):
    usuario_id = call.from_user.id
    print(f"Datos de temp_data en respuesta_zona_dolor: {temp_data}")
    temp_data[usuario_id]['zona_dolor'] = call.data.replace("zona_", "")
    preguntar_intensidad_dolor(call.message)

# Pregunta sobre la intensidad del dolor
def preguntar_intensidad_dolor(message):
    markup = InlineKeyboardMarkup()
    buttons = [InlineKeyboardButton(str(i), callback_data=f"intensidad_{i}") for i in range(1, 11)]
    markup.add(*buttons)
    bot.send_message(message.chat.id, "¿Con qué intensidad se presenta el dolor? (1 = leve, 10 = El peor dolor de todos)", reply_markup=markup)

# Manejar respuestas sobre la intensidad del dolor
@bot.callback_query_handler(func=lambda call: call.data.startswith("intensidad_"))
def respuesta_intensidad_dolor(call):
    usuario_id = call.from_user.id
    print(f"Datos de temp_data en respuesta_intensidad_dolor: {temp_data}")
    temp_data[usuario_id]['intensidad_dolor'] = int(call.data.replace("intensidad_", ""))
    guardar_seguimiento(call.message)

#Guardar seguimiento en la BD
def guardar_seguimiento(message):
    global usuario_id
    #usuario_id = message.from_user.id if message.from_user.id else None

    print(f"Usuario_id: {usuario_id}")
    print(f"temp_data: {temp_data}")
    # Verificar si se obtuvo un usuario_id válido
    if usuario_id is None:
        bot.send_message(message.chat.id, "Error: No se pudo identificar al usuario.")
        return

    # Verificar si el usuario_id existe en temp_data
    if usuario_id not in temp_data:
        bot.send_message(message.chat.id, "Ocurrió un error. No se encontraron los datos de seguimiento.")
        return

    seguimiento = temp_data[usuario_id]
    folio = seguimiento['folio']
    fecha = datetime.now().strftime('%Y-%m-%d')
    hora = datetime.now().strftime('%H:%M:%S')
    temperatura = seguimiento['temperatura']
    vomitos = seguimiento.get('vomitos', '')
    frecuencia_vomitos = seguimiento.get('frecuencia_vomitos', '')
    problemas_respiracion = seguimiento.get('problemas_respiracion', '')
    dolor_corporal = seguimiento.get('dolor_corporal', '')
    zona_dolor = seguimiento.get('zona_dolor', '')
    intensidad_dolor = seguimiento.get('intensidad_dolor', None)

    if not temperatura:
        bot.send_message(message.chat.id, "Error: Falta la temperatura del paciente.")
        print("Error: Falta la temperatura")
        return

    if not vomitos:
        bot.send_message(message.chat.id, "Error: Falta la vomitos del paciente.")
        print("Error: Falta la vomitos")
        return

    if not frecuencia_vomitos:
        bot.send_message(message.chat.id, "Error: Falta la frecuencia vomitos del paciente.")
        print("Error: Falta la frecuencia vomitos")
        return

    if not problemas_respiracion:
        bot.send_message(message.chat.id, "Error:  problemas respiracion del paciente.")
        print("Error: Faltan problemas respiracion")
        return

    if not dolor_corporal:
        bot.send_message(message.chat.id, "Error: Falta la dolor_corporal del paciente.")
        print("Error: Falta la dolor_corporal")
        return

    if not zona_dolor:
        bot.send_message(message.chat.id, "Error: zona_dolor del paciente.")
        print("Error: zona_dolor")
        return

    if not intensidad_dolor:
        bot.send_message(message.chat.id, "Error: Falta la intensidad_dolor del paciente.")
        print("Error: Falta la intensidad_dolor")
        return
    
    # Insertar los datos en la tabla seguimientos
    conn = sqlite3.connect('pacientes.db')
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO seguimientos (folio, fecha, hora, temperatura, vomitos, frecuencia_vomitos, problemas_respiracion, dolor_corporal, zona_dolor, intensidad_dolor)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (folio, fecha, hora, temperatura, vomitos, frecuencia_vomitos, problemas_respiracion, dolor_corporal, zona_dolor, intensidad_dolor))

    conn.commit()
    conn.close()

    bot.send_message(message.chat.id, "Seguimiento guardado con éxito.")

################## BUSCAR SEGUIMIENTO  ##################
def buscar_seguimiento(message):
    markup = InlineKeyboardMarkup()
    buttons = [
        InlineKeyboardButton("Buscar por folio", callback_data="buscar_folio"),
        InlineKeyboardButton("Buscar por fecha", callback_data="buscar_fecha"),
    ]
    markup.add(*buttons)
    bot.send_message(message.chat.id, "Elija una opción para buscar:", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("buscar_"))
def buscar_seguimiento_opciones(call):
    criterio = call.data
    if not criterio:
            bot.reply_to(call, "El criterio de búsqueda no puede estar vacío.")
            bot.reply_to(call, "Ingresa el folio del paciente para buscarlo.")
            bot.register_next_step_handler(call, buscar_seguimiento)
            
    if call.data == "buscar_folio":
        # Solicitar el folio al usuario
        bot.send_message(call.message.chat.id, "Por favor, ingrese el folio del paciente:")
        bot.register_next_step_handler(call.message, procesar_busqueda_folio)
    elif call.data == "buscar_fecha":
        # Solicitar la fecha al usuario
        bot.send_message(call.message.chat.id, "Ingrese folio de paciente.")
        bot.register_next_step_handler(call.message, solicitar_fecha_inicial)

################## BUSCAR SEGUIMIENTO POR FOLIO  ##################
def procesar_busqueda_folio(message):
    folio = message.text  # Aquí obtenemos el texto enviado por el usuario

    if not folio:
        bot.reply_to(message, "El folio no puede estar vacío. Intenta nuevamente.")
        return
    
    # Conectar a la base de datos y buscar por folio
    conn = sqlite3.connect('pacientes.db')
    cursor = conn.cursor()

    try:
        cursor.execute('SELECT * FROM seguimientos WHERE folio = ?', (folio,))
        resultados = cursor.fetchall()
    finally:
        conn.close()

    if not resultados:
        bot.reply_to(message, f"No se encontraron resultados para el folio: {folio}")
        return

    # Crear un nuevo archivo Excel
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Seguimientos_folio"

    # Escribir encabezados
    encabezados = ['Indice', 'Folio', 'Fecha', 'Hora', 'Temperatura', 'Vomitos',
                   'frecuencia_vomitos', 'problemas_respiracion', 'dolor_corporal',
                   'zona_dolor', 'intensidad_dolor']
    for col, encabezado in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=col).value = encabezado
        hoja.cell(row=1, column=col).font = Font(bold=True)

    # Escribir los datos en el archivo
    for fila, datos in enumerate(resultados, start=2):
        for col, valor in enumerate(datos, start=1):
            hoja.cell(row=fila, column=col).value = valor

    # Guardar el archivo temporalmente
    archivo_excel = "reporte_seguimiento.xlsx"
    workbook.save(archivo_excel)

    # Enviar el archivo al usuario
    with open(archivo_excel, 'rb') as file:
        bot.send_document(message.chat.id, file)

    # Eliminar el archivo después de enviarlo
    os.remove(archivo_excel)

################## BUSCAR SEGUIMIENTO POR FECHA  ##################
def solicitar_fecha_inicial(message):
    folio = message.text.strip()
    if not folio:
        bot.reply_to(message, "El folio no puede estar vacío. Intenta nuevamente.")
        return

    seguimiento_data['folio'] = folio  # Guardamos el folio en los datos del chat
    print(seguimiento_data)
    bot.send_message(message.chat.id, "Por favor, ingrese la **fecha inicial** (formato: YYYY-MM-DD):")
    bot.register_next_step_handler(message, solicitar_fecha_final)

def solicitar_fecha_final(message):
    fecha_inicial = message.text.strip()
    try:
        # Validar la fecha inicial
        datetime.strptime(fecha_inicial, "%Y-%m-%d")
    except ValueError:
        bot.reply_to(message, "Fecha inicial no válida. Intenta nuevamente (formato: YYYY-MM-DD).")
        return
    
    seguimiento_data['fecha_inicial'] = fecha_inicial #Guardamos fecha inicial
    bot.send_message(message.chat.id, "Por favor, ingrese la **fecha final** (formato: YYYY-MM-DD):")
    bot.register_next_step_handler(message, procesar_busqueda_fecha)

def procesar_busqueda_fecha(message):
    fecha_final = message.text.strip()
    seguimiento_data['fecha_final'] = fecha_final
    try:
        # Validar la fecha final
        datetime.strptime(fecha_final, "%Y-%m-%d")
    except ValueError:
        bot.reply_to(message, "Fecha final no válida. Intenta nuevamente (formato: YYYY-MM-DD).")
        return

    # Recuperar datos almacenados
    folio = seguimiento_data['folio']
    fecha_inicial = seguimiento_data['fecha_inicial']
    fecha_final = seguimiento_data['fecha_final']

    # Conectar a la base de datos y buscar por rango de fechas
    conn = sqlite3.connect('pacientes.db')
    cursor = conn.cursor()

    try:
        query = '''
        SELECT * FROM seguimientos 
        WHERE folio = ? AND fecha BETWEEN ? AND ?
        '''
        cursor.execute(query, (folio, fecha_inicial, fecha_final))
        resultados = cursor.fetchall()
    finally:
        conn.close()

    if not resultados:
        bot.reply_to(message, f"No se encontraron resultados para el folio '{folio}' entre {fecha_inicial} y {fecha_final}.")
        return

    # Crear un nuevo archivo Excel
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Seguimientos_fecha"

    # Escribir encabezados
    encabezados = ['Indice', 'Folio', 'Fecha', 'Hora', 'Temperatura', 'Vomitos',
                   'frecuencia_vomitos', 'problemas_respiracion', 'dolor_corporal',
                   'zona_dolor', 'intensidad_dolor']
    for col, encabezado in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=col).value = encabezado
        hoja.cell(row=1, column=col).font = Font(bold=True)

    # Escribir los datos en el archivo
    for fila, datos in enumerate(resultados, start=2):
        for col, valor in enumerate(datos, start=1):
            hoja.cell(row=fila, column=col).value = valor

    # Guardar el archivo temporalmente
    archivo_excel = "reporte_seguimiento_fecha.xlsx"
    workbook.save(archivo_excel)

    # Enviar el archivo al usuario
    with open(archivo_excel, 'rb') as file:
        bot.send_document(message.chat.id, file)

    # Eliminar el archivo después de enviarlo
    os.remove(archivo_excel)

################## FUNCIONES UTILES ##################
#Verifica que un paciente se encuentre en la tabla pacientes
def verificar_paciente(folio):
    regex = r"^100+[0-9]{5}"
    conn = sqlite3.connect('pacientes.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM pacientes WHERE folio = ?', (folio,))
    paciente = cursor.fetchone()
    conn.close() # Cerramos la conexión

    if re.fullmatch(regex, folio):
        print("Funciona")
        return paciente
    else:
        bot.reply_to(folio, "El folio no cumple con el formato esperado. Debe comenzar con '100' seguido de 5 números.")
        return
    
#Valida los mensajes que no quiero recibir
@bot.message_handler(func=lambda message: True)
def manejar_mensajes_no_validos(message):
    bot.send_message(
        message.chat.id, 
        "Por favor selecciona una opción utilizando los botones.", 
        reply_markup=ForceReply(selective=True)
    )

if __name__ == '__main__':
    print('Iniciando')
    bot.infinity_polling()